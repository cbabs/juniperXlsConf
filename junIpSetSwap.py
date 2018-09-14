import openpyxl
import netmiko
import getpass
import re
from openpyxl.xml.constants import MIN_ROW
import time

timestamp = time.ctime().replace(':', '.') # Timestamp for file



def login():
    
    while True:
        sshUser = input('Enter username: ')
        sshPass = getpass.getpass('Enter password: ')
        sshIpAddr = input('Enter firewall IP: ')
        
        
        
        juniperSrx = {
        'device_type': 'juniper_junos',
        'ip':   sshIpAddr,
        'username': sshUser,
        'password': sshPass,
        'port': 22 }          # optional, defaults to 22
        
        
        try:
            netmiko.ConnectHandler(**juniperSrx)
            return juniperSrx
            break
    
        except Exception as e:
            print(e)

        else:
            print('Access denied')
            

# Gets the subnet matching rules
def getMatchPolices(subnet, devHndlr):
        
    net_connect = netmiko.ConnectHandler(**devHndlr) # netmiko handler
    
    #Set command for match the list of cmnds.  
    showSubMatch = 'show | display set | match ' + subnet
    set_commands = ['run set cli screen-length 0', showSubMatch]
        
    #Return all output from session.  Needs to be filtered later. It's not tidy
    
    attemptCnt = 0
    
    while attemptCnt < 3:
        attemptCnt += 1
        try:
            retrnedData = net_connect.send_config_set(set_commands,
                                                       delay_factor=3)
            return retrnedData
        except:
            print("Failed to Connect.  Waiting 10 secs and trying again.")
            print(retrnedData)
    
    # return net_connect.send_config_set(set_commands, delay_factor=3)
        

# Primary
def processList(subnetList, devHndlr):
    
    regxAfterLastSpc = r'\s(\S*$)'
    regxAddrMatch = r'\saddress\s'
    
    for items in subnetList:
        
        f = open("fw-{}-{}.txt".format(devHndlr["ip"], timestamp), "a")
        #f2 = open("fw-No-addr-set-{}-{}.txt".format(devHndlr["ip"], timestamp), "a")
        f.write("\n\n") # New ling for each item in excel
        #f2.write("\n\n")
        
        # Get items from dict
        subnetName = items[0]
        subnetName = subnetName.upper()
        currentSubnet = items[1]
        newSdcSubnet = items[2]
        newNdcSubnet = items[3]
        
        addrBookSdcNam = '{}_{}_SDC '.format(newSdcSubnet, subnetName,
                                                 newSdcSubnet) 
        
        addrBookNdcNam = '{}_{}_NDC '.format(newNdcSubnet, subnetName) 
        
        
        print('\n')
        
        setAddrBookCmd = 'set security zones security-zone\
 Outside address-book address ' 
        
        
        setAddrBookCmdSdc = setAddrBookCmd + addrBookSdcNam + newNdcSubnet
        setAddrBookCmdNdc = setAddrBookCmd + addrBookNdcNam + newNdcSubnet
        
        # Write the new addr objects to file
        print(setAddrBookCmdSdc)
        f.write("\n" + setAddrBookCmdSdc+"\n")
        print(setAddrBookCmdNdc)
        f.write(setAddrBookCmdNdc+"\n")
        
        
        newAddrSetName = 'JVPN_{}_USERS' .format(subnetName)
        newAddrSetName = newAddrSetName.upper()
        
        
        setAddrBookSetCmdSdc = 'set security zones security-zone\
 Outside address-book address-set {}_ADD_SET address {}'.format(newAddrSetName, 
                                                    addrBookSdcNam)
        setAddrBookSetCmdNdc = 'set security zones security-zone\
 Outside address-book address-set {}_ADD_SET address {}'.format(newAddrSetName, 
                                                    addrBookNdcNam)
        
        # write the new sets to the file
        print(setAddrBookSetCmdSdc)
        f.write(setAddrBookCmdSdc+"\n")
        print(setAddrBookSetCmdNdc)
        f.write(setAddrBookCmdNdc+"\n")
        
        
        # Returned data from Juniper
        print("Attempting to connect Juniper FW and get matching rules")
        matchData = getMatchPolices(currentSubnet, devHndlr)
        print("Success. Processing rules and creaating now ones")

        # If no data returned raise error
        if not matchData: raise Exception("No matching data found on firewall")
        
        matchAddSetStr = 'set security zones security-zone Outside address-book address-set'
        
        for line in matchData.splitlines():
            
            # Create conf for 
            if matchAddSetStr in line:
                print("\n" + line)
                print('Becomes:')
                if " address " in line:
                    line = re.sub(regxAddrMatch, " address-set ", line)
                    print("' address ' has been chnaged to ' address-set'")
               
                newSecPolCmd = re.sub(regxAfterLastSpc, ' '+newAddrSetName, line)
                print(newSecPolCmd)
                f.write(newSecPolCmd+"\n")
             
                
            if "set security policies from-zone" in line:
           
            
                
                print("\n" + line)
                print('Becomes:')
                newSecPolCmd = re.sub(regxAfterLastSpc, ' '+newAddrSetName, line)
                print(newSecPolCmd)
                f.write(newSecPolCmd+"\n")

       
        print("\n\nNext IP subnet\n")
        
        # Save and close files
        f.close()
        #f2.close()

def readExcel(bookPath=None):
        
    # instantiate book and sheet
    if not bookPath:
        book = openpyxl.load_workbook('juniperConf.xlsx')
    else:    
        book = openpyxl.load_workbook(bookPath)
    sheet = book.active
    
    #List of IP info lists
    changesList = []

    
    # Iterate over rows
    for rows in sheet.iter_rows(min_row=2, min_col=1, max_col=4):

        rowList = [] # Create a row for each loop execution
        for cell in rows:

                rowList.append(cell.value) # Append each cell value to row list
                continue    
            
        changesList.append(rowList) # Append row list to main list

    return changesList #  nested list with all subnets
    
    
def main():
    juniperSrx = login() # gets a dict of ssh reqs for netmiko
    
    subnetList = readExcel()
    print("")
    
    processList(subnetList, juniperSrx)
    

if __name__ == '__main__':
    main()
